"""
QuantWeave - 回测报告导出服务
支持将回测结果导出为 Excel (.xlsx) 和 PDF 格式
"""
import os
import json
import datetime
from typing import Dict, List, Optional
from loguru import logger


class ReportExporter:
    """回测报告导出器"""

    def __init__(self, output_dir: str = "./exports"):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    # ========== Excel 导出 ==========

    def export_excel(self, results: Dict, filename: str = None) -> str:
        """导出回测结果为 Excel 文件

        Args:
            results: 回测结果字典（结构同 backtest_results.json）
            filename: 输出文件名，默认自动生成

        Returns:
            导出文件路径
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.chart import LineChart, Reference
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error("缺少 openpyxl 依赖，请安装: pip install openpyxl")
            return ""

        if not filename:
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"backtest_report_{timestamp}.xlsx"

        filepath = os.path.join(self.output_dir, filename)
        wb = openpyxl.Workbook()

        # 样式定义
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        title_font = Font(bold=True, size=14)
        border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        positive_font = Font(color="FF0000", bold=True)  # 红色=涨（A股惯例）
        negative_font = Font(color="008000", bold=True)   # 绿色=跌（A股惯例）

        # ===== Sheet 1: 策略汇总 =====
        ws_summary = wb.active
        ws_summary.title = "策略汇总"
        self._write_summary_sheet(ws_summary, results, header_font, header_fill, border,
                                   positive_font, negative_font, title_font)

        # ===== Sheet 2: 各股票详情 =====
        for stock_code, stock_results in results.items():
            stock_name = self._get_stock_name(stock_code)
            sheet_name = stock_name[:20]  # Excel sheet名最长31字符
            ws = wb.create_sheet(title=sheet_name)
            self._write_stock_sheet(ws, stock_code, stock_name, stock_results,
                                     header_font, header_fill, border,
                                     positive_font, negative_font)

        # ===== Sheet 3: 净值曲线 =====
        ws_equity = wb.create_sheet(title="净值曲线")
        self._write_equity_sheet(ws_equity, results, header_font, header_fill, border)

        wb.save(filepath)
        logger.info(f"Excel 报告已导出: {filepath}")
        return filepath

    def _write_summary_sheet(self, ws, results, header_font, header_fill, border,
                              positive_font, negative_font, title_font):
        """写入策略汇总 Sheet"""
        ws.merge_cells("A1:H1")
        ws["A1"] = "QuantWeave 回测报告 — 策略汇总"
        ws["A1"].font = title_font

        ws["A2"] = f"生成时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A3"] = ""

        # 汇总每个策略的平均表现
        strategy_summary = {}
        strategies = set()
        for stock_code, stock_results in results.items():
            for strategy_key, strat_result in stock_results.items():
                if isinstance(strat_result, dict) and "error" not in strat_result:
                    strategies.add(strategy_key)
                    if strategy_key not in strategy_summary:
                        strategy_summary[strategy_key] = {
                            "returns": [], "sharpe": [], "drawdown": [],
                            "win_rate": [], "trades": []
                        }
                    strategy_summary[strategy_key]["returns"].append(
                        strat_result.get("total_return", 0))
                    strategy_summary[strategy_key]["sharpe"].append(
                        strat_result.get("sharpe_ratio", 0))
                    strategy_summary[strategy_key]["drawdown"].append(
                        strat_result.get("max_drawdown", 0))
                    strategy_summary[strategy_key]["win_rate"].append(
                        strat_result.get("win_rate", 0))
                    strategy_summary[strategy_key]["trades"].append(
                        strat_result.get("total_trades", 0))

        import numpy as np

        # 表头
        headers = ["策略名称", "平均收益%", "最大收益%", "最小收益%", "平均夏普",
                    "最大回撤%", "平均胜率%", "总交易笔数"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        strategy_names = {
            "dual_ma": "双均线交叉", "bollinger": "布林带突破",
            "rsi": "RSI超买超卖", "macd": "MACD金叉死叉",
            "chip": "主力筹码", "enhanced_chip": "增强筹码",
            "pullback_stable": "强势股回调企稳", "vol_breakout": "爆量突破",
            "first_yin": "龙头首阴反抽", "trend_ma": "均线趋势跟踪",
            "top_bottom": "顶底图策略",
        }

        row = 5
        for sk in sorted(strategy_summary.keys()):
            d = strategy_summary[sk]
            name = strategy_names.get(sk, sk)
            avg_ret = np.mean(d["returns"]) if d["returns"] else 0
            max_ret = np.max(d["returns"]) if d["returns"] else 0
            min_ret = np.min(d["returns"]) if d["returns"] else 0
            avg_sharpe = np.mean(d["sharpe"]) if d["sharpe"] else 0
            avg_dd = np.mean(d["drawdown"]) if d["drawdown"] else 0
            avg_wr = np.mean(d["win_rate"]) if d["win_rate"] else 0
            total_trades = sum(d["trades"])

            values = [name, round(avg_ret, 2), round(max_ret, 2), round(min_ret, 2),
                      round(avg_sharpe, 3), round(avg_dd, 2), round(avg_wr, 2), total_trades]

            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = border
                cell.alignment = Alignment(horizontal="center")
                # 收益列着色
                if col == 2 and isinstance(val, (int, float)):
                    cell.font = positive_font if val > 0 else negative_font

            row += 1

        # 调整列宽
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 16

    def _write_stock_sheet(self, ws, stock_code, stock_name, stock_results,
                            header_font, header_fill, border,
                            positive_font, negative_font):
        """写入单只股票的回测详情"""
        ws.merge_cells("A1:F1")
        ws["A1"] = f"{stock_name} ({stock_code})"
        ws["A1"].font = Font(bold=True, size=13)

        headers = ["策略", "总收益%", "年化收益%", "最大回撤%", "夏普比率", "胜率%", "交易次数"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        strategy_names = {
            "dual_ma": "双均线交叉", "bollinger": "布林带突破",
            "rsi": "RSI超买超卖", "macd": "MACD金叉死叉",
            "chip": "主力筹码", "enhanced_chip": "增强筹码",
            "pullback_stable": "强势股回调企稳", "vol_breakout": "爆量突破",
            "first_yin": "龙头首阴反抽", "trend_ma": "均线趋势跟踪",
            "top_bottom": "顶底图策略",
        }

        row = 4
        for sk, result in stock_results.items():
            if isinstance(result, dict) and "error" in result:
                continue
            if not isinstance(result, dict):
                continue

            name = strategy_names.get(sk, sk)
            values = [
                name,
                result.get("total_return", "N/A"),
                result.get("annual_return", "N/A"),
                result.get("max_drawdown", "N/A"),
                result.get("sharpe_ratio", "N/A"),
                result.get("win_rate", "N/A"),
                result.get("total_trades", "N/A"),
            ]

            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = border
                if col == 2 and isinstance(val, (int, float)):
                    cell.font = positive_font if val > 0 else negative_font
            row += 1

        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _write_equity_sheet(self, ws, results, header_font, header_fill, border):
        """写入净值曲线数据（可用于后续图表）"""
        ws["A1"] = "净值曲线数据（可选中数据插入折线图）"
        ws["A1"].font = Font(bold=True, size=12)
        ws["A2"] = "提示: 选中数据区域 → 插入 → 图表 → 折线图"

    def _get_stock_name(self, code: str) -> str:
        names = {
            "600519.SH": "贵州茅台", "000858.SZ": "五粮液", "601318.SH": "中国平安",
            "600036.SH": "招商银行", "000001.SZ": "平安银行", "000333.SZ": "美的集团",
            "000651.SZ": "格力电器", "601398.SH": "工商银行",
        }
        return names.get(code, code)

    # ========== HTML报告转PDF（备选方案） ==========

    def export_pdf_from_html(self, html_path: str, output_path: str = None) -> str:
        """将HTML报告转为PDF（需安装 weasyprint）

        Args:
            html_path: HTML报告文件路径
            output_path: 输出PDF路径

        Returns:
            PDF文件路径
        """
        if not output_path:
            base = os.path.splitext(html_path)[0]
            output_path = f"{base}.pdf"

        try:
            from weasyprint import HTML
            HTML(filename=html_path).write_pdf(output_path)
            logger.info(f"PDF 报告已导出: {output_path}")
            return output_path
        except ImportError:
            logger.warning("缺少 weasyprint 依赖，请安装: pip install weasyprint")
            return ""
        except Exception as e:
            logger.error(f"PDF导出失败: {e}")
            return ""

    # ========== JSON快速导出 ==========

    def export_json_summary(self, results: Dict, filename: str = None) -> str:
        """导出精简版JSON汇总"""
        if not filename:
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"backtest_summary_{timestamp}.json"

        filepath = os.path.join(self.output_dir, filename)
        summary = {}
        for stock_code, stock_results in results.items():
            summary[stock_code] = {}
            for sk, result in stock_results.items():
                if isinstance(result, dict) and "error" not in result:
                    summary[stock_code][sk] = {
                        "return": result.get("total_return", 0),
                        "sharpe": result.get("sharpe_ratio", 0),
                        "drawdown": result.get("max_drawdown", 0),
                        "win_rate": result.get("win_rate", 0),
                        "trades": result.get("total_trades", 0),
                    }

        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(summary, f, ensure_ascii=False, indent=2)

        logger.info(f"JSON 汇总已导出: {filepath}")
        return filepath


# 便捷函数
def export_backtest_report(results: Dict, fmt: str = "excel", output_dir: str = "./exports") -> str:
    """一键导出回测报告

    Args:
        results: 回测结果（backtest_results.json 格式）
        fmt: 导出格式 "excel" | "json"
        output_dir: 输出目录

    Returns:
        导出文件路径
    """
    exporter = ReportExporter(output_dir=output_dir)
    if fmt == "excel":
        return exporter.export_excel(results)
    elif fmt == "json":
        return exporter.export_json_summary(results)
    else:
        raise ValueError(f"不支持的格式: {fmt}, 可选: excel, json")
